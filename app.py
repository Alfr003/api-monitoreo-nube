from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS
from datetime import datetime, timedelta, timezone
from pathlib import Path
import json
import os
from zoneinfo import ZoneInfo
import csv
import io
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

DATA_FILE = Path("datos_actuales.json")
HIST_FILE = Path("historial.jsonl")  # JSON Lines

# -----------------------------
# Config tabla 5 días
# -----------------------------
HORAS_2H = ["02:00","04:00","06:00","08:00","10:00","12:00","14:00","16:00","18:00","20:00","22:00","24:00"]
DOW_ES = ["Lun.", "Mar.", "Mié.", "Jue.", "Vie.", "Sáb.", "Dom."]

def now_utc():
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

def parse_ts(item: dict):
    """
    Intenta parsear timestamp en estos formatos comunes:
    - "YYYY-MM-DD HH:MM:SS"
    - ISO: "YYYY-MM-DDTHH:MM:SS(.micro)" (con o sin Z)
    """
    s = item.get("timestamp") or item.get("ts_server")
    if not s:
        return None

    s = str(s).strip()

    # "YYYY-MM-DD HH:MM:SS"
    try:
        if "T" not in s and len(s) >= 19 and s[10] == " ":
            return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
    except:
        pass

    # ISO (acepta microsegundos y "Z")
    try:
        return datetime.fromisoformat(s.replace("Z", ""))
    except:
        return None

def get_local_tz():
    # En Render define TZ=America/Costa_Rica
    tzname = os.environ.get("TZ", "UTC")
    try:
        return ZoneInfo(tzname)
    except:
        return ZoneInfo("UTC")

def bucket_hora_2h(dt_local: datetime) -> str:
    """
    Agrupa por bloque de 2 horas usando floor.
    00:00-01:59 -> 24:00 (misma fecha, por cómo está tu tabla)
    """
    h2 = (dt_local.hour // 2) * 2
    if h2 == 0:
        return "24:00"
    return f"{h2:02d}:00"

def iter_historial():
    """Itera historial.jsonl línea por línea (sin cargar todo a memoria)."""
    if not HIST_FILE.exists():
        return
    with HIST_FILE.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except:
                continue

def to_local_dt(item):
    tz = get_local_tz()
    dt = parse_ts(item)
    if not dt:
        return None

    # si viene sin tz, asume UTC
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(tz)
    else:
        dt = dt.astimezone(tz)
    return dt

def read_hist_lines(max_lines=200000):
    """
    Lee historial completo pero recorta a las últimas max_lines líneas
    para no reventar memoria.
    """
    if not HIST_FILE.exists():
        return []
    lines = HIST_FILE.read_text(encoding="utf-8").splitlines()
    if len(lines) > max_lines:
        lines = lines[-max_lines:]
    out = []
    for ln in lines:
        ln = ln.strip()
        if not ln:
            continue
        try:
            out.append(json.loads(ln))
        except:
            continue
    return out

def safe_float(x):
    try:
        return float(x)
    except:
        return None

def build_tabla_5dias(zona: str, days: int = 5, max_lines: int = 8000):
    tz = get_local_tz()
    today = datetime.now(tz).date()

    fechas = [today - timedelta(days=(days - 1 - i)) for i in range(days)]  # viejo->nuevo
    fechas_str = [d.isoformat() for d in fechas]

    celdas = {h: [None] * days for h in HORAS_2H}

    if not HIST_FILE.exists():
        return {
            "zona": zona,
            "fechas": fechas_str,
            "dias": [f"{DOW_ES[d.weekday()]} {d.day:02d}" for d in fechas],
            "horas": HORAS_2H,
            "celdas": celdas
        }

    lines = HIST_FILE.read_text(encoding="utf-8").splitlines()
    if len(lines) > max_lines:
        lines = lines[-max_lines:]

    best = {}  # (fecha_str, bucket) -> (dt_local, item)

    for ln in lines:
        if not ln.strip():
            continue
        try:
            item = json.loads(ln)
        except:
            continue

        if item.get("zona", "Z1") != zona:
            continue

        dt = parse_ts(item)
        if not dt:
            continue

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(tz)
        else:
            dt = dt.astimezone(tz)

        fecha_str = dt.date().isoformat()
        if fecha_str not in fechas_str:
            continue

        bucket = bucket_hora_2h(dt)
        key = (fecha_str, bucket)

        prev = best.get(key)
        if (prev is None) or (dt > prev[0]):
            best[key] = (dt, item)

    for (fecha_str, bucket), (dt, item) in best.items():
        d_index = fechas_str.index(fecha_str)
        t = item.get("temperatura")
        h = item.get("humedad")
        if t is None or h is None:
            continue

        celdas[bucket][d_index] = {
            "t": float(t),
            "h": float(h),
            "ts": dt.isoformat()
        }

    return {
        "zona": zona,
        "fechas": fechas_str,
        "dias": [f"{DOW_ES[d.weekday()]} {d.day:02d}" for d in fechas],
        "horas": HORAS_2H,
        "celdas": celdas
    }

# -----------------------------
# Rutas base
# -----------------------------
@app.get("/")
def home():
    return "API OK. Usa /api/datos, /api/historial, /api/historicos, /api/historial_filtrado, /api/export.xlsx"

@app.get("/api/datos")
def get_datos():
    if DATA_FILE.exists():
        return jsonify(json.loads(DATA_FILE.read_text(encoding="utf-8")))
    return jsonify({"status": "sin_datos"}), 404

@app.get("/api/historial")
def get_historial():
    n = int(request.args.get("n", "200"))
    if not HIST_FILE.exists():
        return jsonify([])
    lines = HIST_FILE.read_text(encoding="utf-8").splitlines()
    data = [json.loads(x) for x in lines[-n:] if x.strip()]
    return jsonify(data)

# -----------------------------
# Históricos tabla 5 días (tu formato actual)
# -----------------------------
@app.get("/api/historicos")
def get_historicos_5dias():
    zona = request.args.get("zona", "Z1")
    return jsonify(build_tabla_5dias(zona=zona, days=5))

# -----------------------------
# ✅ NUEVO: Historial filtrado para tabla (mes/día/hora)
# -----------------------------
@app.get("/api/historial_filtrado")
def historial_filtrado():
    """
    Filtros:
      - zona=Z1
      - month=YYYY-MM      (opcional)
      - date=YYYY-MM-DD    (opcional)
      - hour=HH:MM         (opcional)
      - n=5000             (límite opcional)
    Devuelve lista en orden viejo->nuevo para que la tabla sea cronológica.
    """
    zona = request.args.get("zona", "Z1")
    month = request.args.get("month")   # "2026-01"
    date = request.args.get("date")     # "2026-01-28"
    hour = request.args.get("hour")     # "11:19"
    n = int(request.args.get("n", "5000"))

    data = read_hist_lines(max_lines=200000)
    # zona
    data = [x for x in data if (x.get("zona", "Z1") == zona)]

    def ts_str(item):
        return str(item.get("ts_server") or item.get("timestamp") or "")

    # filtros por prefijo
    if month:
        data = [x for x in data if ts_str(x).startswith(month)]
    if date:
        data = [x for x in data if ts_str(x).startswith(date)]
    if hour:
        # compara HH:MM con slice 11:16
        def match_hour(item):
            s = ts_str(item)
            return len(s) >= 16 and s[11:16] == hour
        data = [x for x in data if match_hour(x)]

    # recorta a los últimos n
    if len(data) > n:
        data = data[-n:]

    # Devuelve SOLO lo necesario para la tabla, en orden viejo->nuevo
    out = []
    for item in data:
        s = ts_str(item)
        fecha = s[:10] if len(s) >= 10 else ""
        hora2 = s[11:16] if len(s) >= 16 else ""

        out.append({
            "fecha": fecha,
            "hora": hora2,
            "temperatura": item.get("temperatura"),
            "humedad": item.get("humedad"),
            "zona": item.get("zona", "Z1"),
            "timestamp": s
        })

    return jsonify(out)

# -----------------------------
# ✅ NUEVO: Exportar Excel (todo o por mes)
# -----------------------------
@app.get("/api/export.xlsx")
def export_xlsx():
    """
    Descarga Excel:
      - zona=Z1
      - month=YYYY-MM (opcional) -> si no viene, exporta TODO
    """
    zona = request.args.get("zona", "Z1")
    month = request.args.get("month")  # opcional

    data = read_hist_lines(max_lines=200000)
    data = [x for x in data if (x.get("zona", "Z1") == zona)]

    def ts(item):
        return str(item.get("ts_server") or item.get("timestamp") or "")

    if month:
        data = [x for x in data if ts(x).startswith(month)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historicos"

    headers = ["Fecha", "Hora", "Temperatura (°C)", "Humedad (%)", "Timestamp", "Zona"]
    ws.append(headers)

    for item in data:
        ts_str = ts(item)
        fecha = ts_str[:10] if len(ts_str) >= 10 else ""
        hora = ts_str[11:16] if len(ts_str) >= 16 else ""
        t = safe_float(item.get("temperatura"))
        h = safe_float(item.get("humedad"))

        ws.append([fecha, hora, t, h, ts_str, item.get("zona", "Z1")])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    nombre = f"historicos_{zona}.xlsx" if not month else f"historicos_{zona}_{month}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre
    )

# -----------------------------
# (Tu export CSV actual - lo dejo por si lo usas)
# -----------------------------
@app.get("/api/historial_export")
def historial_export():
    """
    Exporta CSV (Excel compatible). Params:
      - zona=Z1
      - mes=YYYY-MM (opcional)
    Si no mandas mes => exporta TODO.
    """
    zona = request.args.get("zona", "Z1")
    mes = request.args.get("mes")  # opcional

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["fecha", "hora", "temperatura", "humedad", "zona"])

    for item in iter_historial():
        if item.get("zona", "Z1") != zona:
            continue
        dt = to_local_dt(item)
        if not dt:
            continue
        if mes and dt.strftime("%Y-%m") != mes:
            continue

        writer.writerow([
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M"),
            item.get("temperatura", ""),
            item.get("humedad", ""),
            item.get("zona", "Z1")
        ])

    filename = f"historial_{zona}_{mes if mes else 'TODO'}.csv"
    resp = Response(output.getvalue(), mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# -----------------------------
# POST datos (tu ruta principal)
# -----------------------------
@app.post("/api/datos")
def post_datos():
    # ✅ Seguridad: si existe API_KEY en Render, exige header X-API-KEY
    api_key = os.environ.get("API_KEY")
    if api_key:
        incoming = request.headers.get("X-API-KEY", "")
        if incoming != api_key:
            return jsonify({"status": "forbidden"}), 403

    data = request.get_json(force=True)

    data.setdefault("zona", "Z1")
    data.setdefault("timestamp", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
    data["ts_server"] = now_utc()

    DATA_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    with HIST_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(data, ensure_ascii=False) + "\n")

    return jsonify({"status": "ok"})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
